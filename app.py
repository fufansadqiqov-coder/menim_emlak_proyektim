import streamlit as st
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook, Workbook
import re
import os
import pandas as pd
import time
from rapidfuzz import process, fuzz
from geopy.distance import geodesic
import smtplib
from email.message import EmailMessage
import boto3
from io import BytesIO

# -------------------------------
# 1️⃣ AWS S3 CONFIGURATION
# -------------------------------
import streamlit as st

AWS_ACCESS_KEY = st.secrets["AWS_ACCESS_KEY_ID"]
AWS_SECRET_KEY = st.secrets["AWS_SECRET_ACCESS_KEY"]
S3_BUCKET = "emlak-bot"
S3_FILE_KEY = "HomeSalesData.xlsx"

s3 = boto3.client(
    "s3",
    aws_access_key_id=AWS_ACCESS_KEY,
    aws_secret_access_key=AWS_SECRET_KEY
)

# -------------------------------
# 2️⃣ STREAMLIT CONFIG
# -------------------------------
st.set_page_config(page_title="🏠 Real Estate Web Scraper S3", page_icon="🏘️", layout="centered")
st.title("🏘️ Web Scraper + AWS S3 + Email Sender")

receiver_email = st.text_input("Enter your email to receive the report:")
start_btn = st.button("Start Scraping & Send Email")

# -------------------------------
# 3️⃣ HELPER FUNCTIONS
# -------------------------------

def download_excel_from_s3():
    try:
        obj = s3.get_object(Bucket=S3_BUCKET, Key=S3_FILE_KEY)
        data = obj['Body'].read()
        wb = load_workbook(filename=BytesIO(data))
        return wb
    except s3.exceptions.NoSuchKey:
        st.info("S3 Excel not found. A new workbook will be created.")
        wb = Workbook()
        ws = wb.active
        ws.append(["Qiymet","Erazi","Link","Lat","Lng","Elan yerlesdirilme tarixi",
                     "Kateqoriya","Mərtəbə	Sahə","Otaq sayı","Çıxarıs","Təmir","İpoteka"])
        return wb

def upload_excel_to_s3(wb):
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    s3.put_object(Bucket=S3_BUCKET, Key=S3_FILE_KEY, Body=buffer.getvalue())
    st.success("✅ Excel uploaded to S3 successfully.")

def send_email(sender, password, receiver, subject, body):
    msg = EmailMessage()
    msg['Subject'] = subject
    msg['From'] = sender
    msg['To'] = receiver
    msg.set_content(body)
    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
            smtp.login(sender, password)
            smtp.send_message(msg)
        return True
    except Exception as e:
        st.error(f"Email send error: {e}")
        return False

# -------------------------------
# 4️⃣ SCRAPING FUNCTION
# -------------------------------
def run_scraping(receiver_email):
    st.info("Starting scraping... ⏳")
    
    # Load Excel from S3
    wb = download_excel_from_s3()
    ws = wb.active
    existing_headers = [cell.value for cell in ws[1]]
    
    # Demo scraping URLs
    HEADERS = {"User-Agent": "Mozilla/5.0"}
    yeni_fursetler = []

    for page in range(1,3):
        url = f"https://bina.az/items/5544660?page={page}"
        st.write(f"Scraping page {page} ...")
        try:
            r = requests.get(url, headers=HEADERS)
            r.raise_for_status()
        except:
            continue

        soup = BeautifulSoup(r.text, "html.parser")
        base_url = "https://bina.az"

        links = set(a.get("href") for a in soup.find_all("a", class_="item_link") if a.get("href"))
        for href in list(links)[:5]:
            link = base_url + href
            try:
                r_item = requests.get(link, headers=HEADERS)
                r_item.raise_for_status()
            except:
                continue

            soup_item = BeautifulSoup(r_item.text, "html.parser")
            try:
                map_div = soup_item.find("div", {"id":"item_map"})
                konum = soup_item.find("h1", class_="product-title").get_text(strip=True)
                elan_tarixi = None
                stat_elements = soup_item.find_all("span", class_="product-statistics__i-text")
                if len(stat_elements) > 1:
                    match = re.search(r"(\d{2}\.\d{2}\.\d{4}), (\d{2}:\d{2})", stat_elements[1].get_text())
                    elan_tarixi = match.group() if match else None

                melumat = {
                    "Qiymet": soup_item.find("div", class_="product-price__i").get_text(strip=True),
                    "Erazi": konum.split(",")[-1].strip(),
                    "Link": link,
                    "Lat": map_div.get("data-lat") if map_div else None,
                    "Lng": map_div.get("data-lng") if map_div else None,
                    "Elan yerlesdirilme tarixi": elan_tarixi
                }


                # Add new columns if they don't exist
                for col in melumat.keys():
                    if col not in existing_headers:
                        existing_headers.append(col)
                        ws.cell(row=1, column=len(existing_headers), value=col)

                row_data = [melumat.get(col, "") for col in existing_headers]
                ws.append(row_data)
                yeni_fursetler.append(f"🏠 {melumat['Erazi']} → {melumat['Qiymet']}")
            except:
                continue
            time.sleep(0.2)

    # Upload Excel back to S3
    upload_excel_to_s3(wb)

    # Send email
    if yeni_fursetler:
        body = "\n".join(yeni_fursetler)
        if send_email("fufansadqiqov@gmail.com","ygug qfyf vylb hsxb", receiver_email, "Yeni daşınmaz əmlak fürsətləri 💰", body):
            st.success(f"📧 Email sent to {receiver_email}")
    else:
        st.warning("No new properties found.")

# -------------------------------
# 5️⃣ TRIGGER
# -------------------------------
if start_btn:
    if receiver_email:
        run_scraping(receiver_email)
    else:
        st.warning("Please enter your email first.")



