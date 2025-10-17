import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import re
import pandas as pd
import time
from rapidfuzz import process, fuzz
from geopy.distance import geodesic
import smtplib
from email.message import EmailMessage

# Data
Mekan_qiymetleri = pd.read_excel(r"C:\Users\hp\Downloads\HomeSales 2.xlsx")
# metrodata
metro_data = pd.read_csv(r"C:\Users\hp\Downloads\Stansiya long lat.csv")

# Kvadrat metrə düşən qiymət (təkrar hissə silindi)
Mekan_qiymetleri["Qiymet"] = Mekan_qiymetleri["Qiymet"].astype(str).str.extract(r"(\d*\s*\d*)")[0].str.replace(" ","").astype("int64")
Mekan_qiymetleri["Sahə"] = Mekan_qiymetleri["Sahə"].astype(str).str.extract(r"(\d*)")[0].astype("int64")

Mekan_qiymetleri["Qiymet_m2"] = Mekan_qiymetleri["Qiymet"] / Mekan_qiymetleri["Sahə"]

# Hər ərazi üçün ortalama Qiymet/m²
ortalama_m2 = Mekan_qiymetleri.groupby("Erazi")["Qiymet_m2"].median().reset_index()
ortalama_m2.rename(columns={"Qiymet_m2": "Ortalama_Qiymet_m2"}, inplace=True)

# --- KONFİQURASİYA ---
HEADERS = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}
excel_path = r"C:\Users\hp\HomeSalesData.xlsx"
sender_email = "fufansadqiqov@gmail.com"
app_password = "ygug qfyf vylb hsxb"
receiver_email = ["ugurnihat123321@gmail.com", "eyvazazim@gmail.com"]
SLEEP_TIME = 0.3

# --- NORMALIZASİYA ---
def normalize_area(area, standard_areas):
    area = area.lower().strip()
    result = process.extractOne(area, standard_areas, scorer=fuzz.partial_ratio)
    if result:
        best_match, score = result[0], result[1]
        if score > 70:
            return best_match
    return area

# --- Qiymət/m² Transform funksiyası ---
def transform_for_stats(melumat, standard_areas):
    mekan_raw = melumat["Erazi"].split(",")[-1].strip()
    mekan = normalize_area(mekan_raw, standard_areas)

    match = re.search(r"([\d\s]+)([A-Za-z]+)", melumat["Qiymet"])
    if match:
        qiymet, valyuta = match.groups()
        qiymet = int(qiymet.replace(" ", ""))
    else:
        qiymet, valyuta = None, None

    sahe_match = re.search(r"(\d+)", melumat.get("Sahə", ""))
    sahe = int(sahe_match.group(1)) if sahe_match else None

    qiymet_m2 = qiymet / sahe if qiymet and sahe else None

    return {
        "Mekan": mekan,
        "Qiymet": qiymet,
        "Valyuta": valyuta,
        "Sahe": sahe,
        "Qiymet_m2": qiymet_m2,
        "Link": melumat.get("Link", ""),
        "Lat": melumat.get("Lat", None),
        "Lng": melumat.get("Lng", None),
        "Elan yerlesdirilme tarixi": melumat.get("Elan yerlesdirilme tarixi", None)
    }

# --- Ən yaxın metro hesabı ---
def nearest_metro_distance(lat, lng, metro_data):
    try:
        house_point = (float(lat), float(lng))
        metro_data["Mesafe_m"] = metro_data.apply(
            lambda row: geodesic(house_point, (row["Enlem"], row["Boylam"])).meters,
            axis=1
        )
        nearest = metro_data.loc[metro_data["Mesafe_m"].idxmin()]
        return nearest["İstasyon Adı"], nearest["Mesafe_m"]
    except:
        return None, None

# # --- Email göndərmə ---
# def send_email(sender, password, receivers, subject, body):
#     msg = EmailMessage()
#     msg['Subject'] = subject
#     msg['From'] = sender
#     msg['To'] = ", ".join(receivers)
#     msg.set_content(body)

#     with smtplib.SMTP("smtp.gmail.com", 587) as server:
#         server.starttls()
#         server.login(sender, password)
#         server.send_message(msg)

# --- Əsas hissə ---
standard_areas = Mekan_qiymetleri["Erazi"].str.lower().tolist()
yeni_fursetler = []

wb = load_workbook(excel_path)
ws = wb.active
existing_headers = [cell.value for cell in ws[1]]

for i in range(1, 3):
    print("page:", i)
    url = f"https://bina.az/items/5544660?page={i}"
    
    try:
        r = requests.get(url, headers=HEADERS, timeout=10)
        r.raise_for_status()
    except Exception as e:
        print(f"Səhifə xətası: {e}")
        continue
        
    soup = BeautifulSoup(r.text, "html.parser")
    base_url = "https://bina.az"

    linkler = set()
    for a in soup.find_all("a", class_="item_link"):
        href = a.get("href")
        if href and href.startswith("/items/"):
            linkler.add(href)

    print(f"Səhifə {i} üçün tapılan link sayı: {len(linkler)}")

    for href in list(linkler)[:5]:
        link = base_url + href
        print(f"İşlənir: {link}")

        try:
            r_item = requests.get(link, headers=HEADERS, timeout=10)
            r_item.raise_for_status()
        except Exception as e:
            print(f"⚠️ Xəta: {e} → keçilir: {link}")
            continue

        soup_item = BeautifulSoup(r_item.text, "html.parser")
        
        # Məlumatları çıxar
        try:
            map_div = soup_item.find("div", {"id": "item_map"})
            konum = soup_item.find("h1", class_="product-title").get_text(strip=True)
            
            # Elan tarixi
            stat_elements = soup_item.find_all("span", class_="product-statistics__i-text")
            elan_tarixi_str = None
            if len(stat_elements) > 1:
                elan_tarixi_text = stat_elements[1].get_text()
                match = re.search(r'(\d{2}\.\d{2}\.\d{4}), (\d{2}:\d{2})', elan_tarixi_text)
                elan_tarixi_str = match.group() if match else None

            melumat = {
                "Qiymet": soup_item.find("div", class_="product-price__i").get_text(strip=True),
                "Erazi": konum.split(",")[-1].replace("\xa0", " ").strip(),
                "Link": link,
                "Lat": map_div.get("data-lat") if map_div else None,
                "Lng": map_div.get("data-lng") if map_div else None,
                "Elan yerlesdirilme tarixi": elan_tarixi_str
            }

            # Əlavə məlumatlar
            kategori = [i.get_text(strip=True) for i in soup_item.find_all("label", class_="product-properties__i-name")]
            deyer = [k.get_text(strip=True) for k in soup_item.find_all("span", class_="product-properties__i-value")]
            
            for idx in range(min(len(kategori), len(deyer))):
                melumat[kategori[idx]] = deyer[idx]

        except Exception as e:
            print(f"Məlumat çıxarılması xətası: {e}")
            continue

        # Excel əməliyyatları
        new_cols = [k for k in melumat.keys() if k not in existing_headers]
        for col in new_cols:
            ws.cell(row=1, column=len(existing_headers)+1, value=col)
            existing_headers.append(col)

        row = [melumat.get(col, "") for col in existing_headers]
        ws.append(row)

        # Transform + analitik
        transformed = transform_for_stats(melumat, standard_areas)
        print(f"Transform edildi: {transformed['Mekan']} - Qiymət: {transformed['Qiymet_m2']}")

        # Metro məsafəsi
        if transformed["Lat"] and transformed["Lng"]:
            nearest_metro, mesafe_m = nearest_metro_distance(transformed["Lat"], transformed["Lng"], metro_data)
        else:
            nearest_metro, mesafe_m = None, None

        # Qiymət müqayisəsi
        if transformed["Sahe"] and transformed["Qiymet_m2"]:
            df_area = ortalama_m2[ortalama_m2["Erazi"].str.lower() == transformed["Mekan"].lower()]
            if not df_area.empty:
                ortalama_m2_value = df_area["Ortalama_Qiymet_m2"].values[0]
                print(f"Müqayisə: {transformed['Qiymet_m2']:.2f} vs {ortalama_m2_value:.2f}")

                if transformed["Qiymet_m2"] < ortalama_m2_value:
                    if mesafe_m and mesafe_m <= 500:
                        msg_title = "🚨 İnanılmaz fürsət!"
                    else:
                        msg_title = "🏠 Yeni fürsət tapıldı!"
                    
                    msg = (
                        f"{msg_title}\n"
                        f"Ərazi: {transformed['Mekan']}\n"
                        f"Qiymət/m²: {transformed['Qiymet_m2']:.2f} AZN\n"
                        f"Ortalama: {ortalama_m2_value:.2f} AZN\n"
                        f"Metro: {nearest_metro} ({mesafe_m/1000:.2f} km məsafədə)\n"
                        f"Link: {transformed['Link']}\n"
                        f"Koordinatlar: {transformed['Lat']}, {transformed['Lng']}\n"
                        f"Elan tarixi: {transformed['Elan yerlesdirilme tarixi']}"
                    )
                    yeni_fursetler.append(msg)
                    print("Yeni fürsət əlavə edildi!")

        time.sleep(SLEEP_TIME)

# Excel save
wb.save(excel_path)
wb.close()

# Mail göndər
if yeni_fursetler:
    mail_body = "\n\n".join(yeni_fursetler)
    send_email(sender_email, app_password, receiver_email, "Yeni daşınmaz əmlak fürsətləri 💰", mail_body)
    print(f"📧 {len(yeni_fursetler)} fürsət ilə mail göndərildi.")
else:
    print("Yeni fürsət tapılmadı.")